#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Excelワークブック作成
wb = Workbook()

# デフォルトシートを削除
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# スタイル定義
header_fill = PatternFill(start_color="01C1AF", end_color="01C1AF", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")
subheader_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_column_widths(ws, widths):
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

def add_header(ws, row, headers, fill, font):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

def add_data_row(ws, row, data):
    for col_num, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col_num, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border

# ==========================================
# シート1: 新人介護士
# ==========================================
ws1 = wb.create_sheet("新人介護士")
ws1.merge_cells('A1:F1')
ws1['A1'] = '新人介護士（入職1年目）の課題解決マップ'
ws1['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

# 朝の業務
ws1.merge_cells('A3:F3')
ws1['A3'] = '📌 朝の業務（7:00～9:00）'
ws1['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws1['A3'].font = Font(bold=True, size=11)
ws1['A3'].alignment = Alignment(horizontal='left', vertical='center')

headers = ['誰が', 'どんな業務・作業で', '何をする', '現場の課題', 'このページで解消できること', '効果']
add_header(ws1, 4, headers, header_fill, header_font)

morning_data = [
    ['新人介護士', '起床介助の準備中に', '確認する', '「岡田さんは何時に起こせばいいんだっけ？」と毎回リーダーに聞く', '左カラムの24時間シートを見れば、07:00起床と一目で分かる', '確認時間ゼロ'],
    ['新人介護士', '岡田さんの部屋に入る前に', '読む', '紙の申し送りノートを探して、前夜の様子を確認するのに5分かかる', '中央カラムのケース記録をスクロールすれば、前夜の記録がすぐ見つかる', '5分→30秒'],
    ['新人介護士', '起床介助中に', '思い出す', '「右側から声をかける、だっけ？左だっけ？」と不安になる', '24時間シートの「詳細」欄に「右側から声をかける」と書いてある', '不安解消'],
    ['新人介護士', '朝食介助の準備中に', '覚える', '「岡田さんはお茶とコーヒー、どっちが好きだっけ？」と迷う', '左カラムに「朝のコーヒーと庭の花を眺める時間が好き」と表示', '迷う時間ゼロ'],
    ['新人介護士', '朝食後に', '記録する', '「今すぐ記録したいけど、記録用紙が事務所にある」と後回しにして忘れる', 'スマホから、その場で「コーヒーを完飲されました」と入力', '記録漏れゼロ'],
]

for i, data in enumerate(morning_data, 5):
    add_data_row(ws1, i, data)

# 日中の業務
row = len(morning_data) + 6
ws1.merge_cells(f'A{row}:F{row}')
ws1[f'A{row}'] = '📌 日中の業務（10:00～17:00）'
ws1[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws1[f'A{row}'].font = Font(bold=True, size=11)
ws1[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws1, row, headers, header_fill, header_font)

daytime_data = [
    ['新人介護士', '排泄介助中に', '判断する', '「岡田さん、自分で立とうとしてるけど、介助していいのかな？」と迷う', '24時間シートに「立位保持を介助」と書いてあるので、迷わず実行', '判断時間ゼロ'],
    ['新人介護士', '午後のレクリエーション準備中に', '気づく', '「岡田さん、最近あまり笑わないな」と感じるが、どう伝えればいいか分からない', '右カラムの「気づきを書く」ボタンで、「最近笑顔が少ない気がします」と投稿', '気づきを共有'],
    ['新人介護士', 'ベテランの介助を見学中に', '学ぶ', '「佐藤リーダーの声かけが上手だな」と思うが、メモする暇がない', '後で「コツのフィット」カテゴリで、「佐藤リーダーの声かけが効果的でした」と投稿', '学習機会増加'],
    ['新人介護士', '入浴介助の準備中に', '確認する', '「岡田さんの入浴は何曜日だっけ？」とシフト表と照合するのが面倒', '24時間シートに「週3回入浴」と書いてあり、過去の記録から曜日も分かる', '確認時間短縮'],
]

row += 1
for i, data in enumerate(daytime_data, row):
    add_data_row(ws1, i, data)

# 夜勤
row = row + len(daytime_data) + 1
ws1.merge_cells(f'A{row}:F{row}')
ws1[f'A{row}'] = '📌 夜勤（17:00～翌9:00）'
ws1[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws1[f'A{row}'].font = Font(bold=True, size=11)
ws1[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws1, row, headers, header_fill, header_font)

night_data = [
    ['新人介護士', '夜間の見回り中に', '記憶する', '「岡田さん、さっき何時にトイレ行ったっけ？」と忘れてしまう', 'その場でスマホから「22:15 トイレ誘導」と記録', '記憶不要'],
    ['新人介護士', '夜勤中に岡田さんが不穏になった時', '悩む', '「どう対応すればいいか分からない。リーダーは昼間しかいない」と孤立する', 'AI相談窓口（将来実装）で「夜間の不穏時の対応は？」と質問', '孤立感解消'],
    ['新人介護士', '朝の申し送り準備中に', 'まとめる', '夜勤中の出来事をメモから思い出して、申し送りノートに清書するのに30分', 'ケース記録に入力済みなので、画面を見せるだけで申し送り完了', '30分→5分'],
]

row += 1
for i, data in enumerate(night_data, row):
    add_data_row(ws1, i, data)

set_column_widths(ws1, [12, 20, 12, 35, 40, 15])

# ==========================================
# シート2: ベテラン介護士
# ==========================================
ws2 = wb.create_sheet("ベテラン介護士")
ws2.merge_cells('A1:F1')
ws2['A1'] = 'ベテラン介護士（経験5年以上）の課題解決マップ'
ws2['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws2.merge_cells('A3:F3')
ws2['A3'] = '📌 朝の業務（7:00～9:00）'
ws2['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws2['A3'].font = Font(bold=True, size=11)
ws2['A3'].alignment = Alignment(horizontal='left', vertical='center')

add_header(ws2, 4, headers, header_fill, header_font)

veteran_data = [
    ['ベテラン介護士', '起床介助中に', '気づく', '「岡田さん、最近8時過ぎまで寝てるな。計画は7時起床だけど…」と変化に気づく', '右カラムで「時間のフィット」を選び、「最近は8:15まで熟睡。時間を遅らせては？」と提案', '気づきを提案に'],
    ['ベテラン介護士', '朝食介助中に', '工夫する', '「岡田さん、食事前に時計を見せると落ち着くんだよな」というコツを持っているが、新人に伝える時間がない', '右カラムで「コツのフィット」を選び、「食事前に時計を見せると拒否が減ります」と投稿', '暗黙知を明文化'],
    ['ベテラン介護士', '朝食後の片付け中に', '比較する', '「最近のお茶の飲み残しが多いな。前はお茶が好きだったのに」と感じる', '右カラムで「好みのフィット」を選び、「お茶よりコーヒーを好まれる様子」と投稿', '変化を記録'],
]

for i, data in enumerate(veteran_data, 5):
    add_data_row(ws2, i, data)

row = 8
ws2.merge_cells(f'A{row}:F{row}')
ws2[f'A{row}'] = '📌 日中の業務（10:00～17:00）'
ws2[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws2[f'A{row}'].font = Font(bold=True, size=11)
ws2[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws2, row, headers, header_fill, header_font)

veteran_daytime = [
    ['ベテラン介護士', '午後のティータイム準備中に', '試す', '「岡田さん、コーヒーが好きみたいだから、今日はコーヒーを出してみよう」と実験する', 'ティータイム後に「コーヒーで笑顔が増えました」と記録。効果を検証', '実験→検証'],
    ['ベテラン介護士', '新人に指導中に', '伝える', '「岡田さんのこだわりは右側から声をかけることなんだよ」と口頭で伝えるが、新人が覚えきれない', '24時間シートの「詳細」欄に書いておけば、新人がいつでも確認できる', '指導時間削減'],
    ['ベテラン介護士', 'カンファレンス準備中に', 'まとめる', '「岡田さんの14日間の変化をまとめて、家族に報告しなきゃ」と資料作成に1時間', 'Day 14で自動生成された「完成版24時間シート」を印刷して、家族に渡すだけ', '1時間→10分'],
]

row += 1
for i, data in enumerate(veteran_daytime, row):
    add_data_row(ws2, i, data)

set_column_widths(ws2, [15, 20, 12, 35, 40, 15])

# ==========================================
# シート3: リーダー・ケアマネ
# ==========================================
ws3 = wb.create_sheet("リーダー・ケアマネ")
ws3.merge_cells('A1:F1')
ws3['A1'] = 'リーダー・ケアマネージャーの課題解決マップ'
ws3['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells('A3:F3')
ws3['A3'] = '📌 朝礼・カンファレンス（8:30～9:00）'
ws3['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws3['A3'].font = Font(bold=True, size=11)
ws3['A3'].alignment = Alignment(horizontal='left', vertical='center')

add_header(ws3, 4, headers, header_fill, header_font)

leader_data = [
    ['リーダー', '朝礼で新人に指示する時', '指示する', '「岡田さんは右側から声をかけてね」と毎朝口頭で伝えるが、言い忘れることも', '新人が24時間シートを見れば分かるので、朝礼時間を短縮できる', '朝礼時間短縮'],
    ['リーダー', 'カンファレンスで提案を検討する時', '判断する', '「岡田さんの起床時刻を遅らせる提案があったけど、根拠が曖昧」と判断に迷う', '右カラムの「気づきのストック」を見れば、「最近8:15まで熟睡」という記録が根拠になる', '判断が明確に'],
    ['ケアマネ', 'ケアプラン作成中に', '収集する', '「岡田さんの最近の様子を知りたい」と、各スタッフに聞いて回るのに30分', 'ケース記録を見れば、全スタッフの記録が時系列で確認できる', '30分→5分'],
]

for i, data in enumerate(leader_data, 5):
    add_data_row(ws3, i, data)

row = 8
ws3.merge_cells(f'A{row}:F{row}')
ws3[f'A{row}'] = '📌 日中の管理業務（10:00～17:00）'
ws3[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws3[f'A{row}'].font = Font(bold=True, size=11)
ws3[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws3, row, headers, header_fill, header_font)

leader_daytime = [
    ['リーダー', 'スタッフの気づきを確認する時', '確認する', '「スタッフから何か提案があったはず」と思うが、口頭で聞いたことを忘れる', '右カラムの「気づきのストック」に、全スタッフの気づきが集約されている', '提案の見逃しゼロ'],
    ['リーダー', '気づきを採用するか判断する時', '決める', '「この提案を採用していいか？」と悩むが、判断基準が曖昧', '付せんの「とりいれる」ボタンを押せば、計画に反映。「まだそのまま」で保留も可能', '判断フロー明確'],
    ['ケアマネ', '家族面談の準備中に', '説明する', '「岡田さんの14日間の様子を家族に説明したい」と、手書きメモをまとめるのに1時間', '完成版24時間シートを印刷して、「このように生活リズムを調整しました」と説明', '1時間→10分'],
    ['リーダー', '新人育成中に', '育てる', '「新人がベテランのコツを学ぶ機会が少ない」と悩む', '「コツのフィット」に蓄積されたベテランの知恵を、新人に読ませる', '育成効率向上'],
]

row += 1
for i, data in enumerate(leader_daytime, row):
    add_data_row(ws3, i, data)

set_column_widths(ws3, [12, 20, 12, 35, 40, 15])

# ==========================================
# シート4: 夜勤専従・パート
# ==========================================
ws4 = wb.create_sheet("夜勤専従・パート")
ws4.merge_cells('A1:F1')
ws4['A1'] = '夜勤専従スタッフ・パート職員の課題解決マップ'
ws4['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws4.merge_cells('A3:F3')
ws4['A3'] = '🏠 夜勤専従スタッフ：夜間巡回（20:00～翌6:00）'
ws4['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws4['A3'].font = Font(bold=True, size=11)
ws4['A3'].alignment = Alignment(horizontal='left', vertical='center')

add_header(ws4, 4, headers, header_fill, header_font)

night_staff_data = [
    ['夜勤専従', '夜勤開始時の申し送りで', '引き継ぐ', '昼のスタッフから口頭で申し送りを受けるが、聞き逃す', 'ケース記録を見れば、今日一日の様子が全て記録されている', '申し送り漏れゼロ'],
    ['夜勤専従', '深夜の見回り中に', '記録する', '「岡田さん、23時にトイレ誘導した」とメモするが、後で清書するのを忘れる', 'その場でスマホから「23:00 トイレ誘導、自力で歩行」と入力', '記録漏れゼロ'],
    ['夜勤専従', '早朝の起床介助前に', '確認する', '「岡田さんの起床時刻、7時だっけ？8時だっけ？」と不安', '24時間シートを見れば、07:00起床と書いてある（最近の変化も付せんで分かる）', '不安解消'],
    ['夜勤専従', '夜勤明けの申し送り準備中に', 'まとめる', '「今夜の出来事をノートにまとめなきゃ」と、手書きで清書するのに30分', 'ケース記録に入力済みなので、口頭で補足するだけで申し送り完了', '30分→5分'],
]

for i, data in enumerate(night_staff_data, 5):
    add_data_row(ws4, i, data)

row = 9
ws4.merge_cells(f'A{row}:F{row}')
ws4[f'A{row}'] = '👪 パート・非常勤スタッフ：週2回勤務の業務（10:00～17:00）'
ws4[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws4[f'A{row}'].font = Font(bold=True, size=11)
ws4[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws4, row, headers, header_fill, header_font)

part_staff_data = [
    ['パートスタッフ', '久しぶりの勤務開始時に', '思い出す', '「岡田さん、前回会ったのは1週間前。今の様子はどうだろう？」と不安', 'ケース記録で、過去7日間の記録を一気に確認できる', '情報不足解消'],
    ['パートスタッフ', '食事介助の準備中に', '確認する', '「岡田さんの食事形態、一口大だっけ？刻みだっけ？」と毎回リーダーに確認', '24時間シートに「主食は一口大。副食は刻み」と書いてある', '確認時間ゼロ'],
    ['パートスタッフ', 'ケア中に工夫を思いついた時', '提案する', '「この方法、良さそう」と思うが、リーダーに伝える機会がない', '右カラムで「コツのフィット」に投稿。リーダーが後で確認してくれる', '提案機会増加'],
]

row += 1
for i, data in enumerate(part_staff_data, row):
    add_data_row(ws4, i, data)

set_column_widths(ws4, [12, 20, 12, 35, 40, 15])

# ==========================================
# シート5: 業務フロー別
# ==========================================
ws5 = wb.create_sheet("業務フロー別")
ws5.merge_cells('A1:F1')
ws5['A1'] = '業務フロー別の課題解決マップ（14日間サイクル）'
ws5['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws5.merge_cells('A3:F3')
ws5['A3'] = '📅 入所時（Day 1）'
ws5['A3'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws5['A3'].font = Font(bold=True, size=11)
ws5['A3'].alignment = Alignment(horizontal='left', vertical='center')

add_header(ws5, 4, headers, header_fill, header_font)

flow_day1 = [
    ['ケアマネ', '24時間シート作成で', '作る', 'アセスメントシートを見ながら、手書きで24時間シートを作成。3時間かかる', 'アセスメントシートをアップロードすれば、AIが暫定シートを自動生成（将来実装）', '3時間→30分'],
    ['リーダー', '新入居者の情報共有で', '伝える', '朝礼で口頭で「岡田さんのこだわり」を全員に伝えるが、覚えきれない', '左カラムに「こだわり」を表示すれば、全員がいつでも確認可能', '伝達漏れゼロ'],
]

for i, data in enumerate(flow_day1, 5):
    add_data_row(ws5, i, data)

row = 7
ws5.merge_cells(f'A{row}:F{row}')
ws5[f'A{row}'] = '📅 実態ログ蓄積期（Day 2-7）'
ws5[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws5[f'A{row}'].font = Font(bold=True, size=11)
ws5[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws5, row, headers, header_fill, header_font)

flow_day27 = [
    ['全スタッフ', '日々のケア記録で', '記録する', '「後で書こう」と思って忘れる。紙に書くのが面倒', 'その場でスマホから入力。時刻・スタッフ名は自動記録', '記録漏れ削減'],
    ['リーダー', '記録の確認で', '確認する', '「みんなちゃんと記録してるかな？」と、紙のノートを確認して回る', 'ケース記録画面で、全員の記録を一覧表示。未記録も分かる', '確認時間短縮'],
]

row += 1
for i, data in enumerate(flow_day27, row):
    add_data_row(ws5, i, data)

row = row + len(flow_day27) + 1
ws5.merge_cells(f'A{row}:F{row}')
ws5[f'A{row}'] = '📅 フィット提案期（Day 8-13）'
ws5[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws5[f'A{row}'].font = Font(bold=True, size=11)
ws5[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws5, row, headers, header_fill, header_font)

flow_day813 = [
    ['ベテラン', '気づきの投稿で', '提案する', '「起床時刻を遅らせた方がいい」と思うが、提案の場がない', '右カラムで「時間のフィット」に投稿。根拠も書ける', '提案が可視化'],
    ['リーダー', '提案の採用判断で', '決める', '「この提案、採用していいかな？」と悩むが、判断材料が少ない', '付せんに記録の根拠が書いてあるので、判断しやすい', '判断が明確に'],
    ['新人', '提案の確認で', '学ぶ', '「ベテランはどんな工夫をしてるんだろう？」と知りたいが、聞けない', '「気づきのストック」に、全員の工夫が蓄積されている', '学習機会増加'],
]

row += 1
for i, data in enumerate(flow_day813, row):
    add_data_row(ws5, i, data)

row = row + len(flow_day813) + 1
ws5.merge_cells(f'A{row}:F{row}')
ws5[f'A{row}'] = '📅 プラン確定（Day 14）'
ws5[f'A{row}'].fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
ws5[f'A{row}'].font = Font(bold=True, size=11)
ws5[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

row += 1
add_header(ws5, row, headers, header_fill, header_font)

flow_day14 = [
    ['ケアマネ', '24時間シート完成で', 'まとめる', '14日間の記録をもとに、手書きで24時間シートを清書。2時間かかる', 'Day 14で自動生成された「完成版シート」を確認・承認するだけ', '2時間→15分'],
    ['リーダー', '家族への報告で', '説明する', '「14日間でどう変わったか」を家族に説明する資料を手作り', '完成版シートを印刷して、「このように調整しました」と説明', '資料作成不要'],
]

row += 1
for i, data in enumerate(flow_day14, row):
    add_data_row(ws5, i, data)

set_column_widths(ws5, [12, 20, 12, 35, 40, 15])

# ==========================================
# シート6: 時間削減効果
# ==========================================
ws6 = wb.create_sheet("時間削減効果")
ws6.merge_cells('A1:F1')
ws6['A1'] = '⏰ 時間削減効果の具体例'
ws6['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws6['A1'].alignment = Alignment(horizontal='center', vertical='center')

time_headers = ['業務', '従来の方法', 'かかる時間', 'このページでの方法', '短縮後の時間', '削減率']
add_header(ws6, 3, time_headers, header_fill, header_font)

time_data = [
    ['申し送り準備', '手書きメモをノートに清書', '30分', 'ケース記録を見せるだけ', '5分', '83%削減'],
    ['24時間シート作成', '手書きで一から作成', '3時間', '暫定シートを自動生成', '30分', '83%削減'],
    ['過去の記録確認', '紙のノートを探す', '10分', 'ケース記録で検索', '30秒', '95%削減'],
    ['気づきの共有', '朝礼で口頭報告', '10分', '付せん投稿', '2分', '80%削減'],
    ['ケア中の確認', 'リーダーに確認の電話', '5分', '24時間シートを見る', '10秒', '97%削減'],
    ['ケアプラン更新', '紙のシートを書き直し', '1時間', '付せんを採用してデジタル更新', '10分', '83%削減'],
    ['新人への指導', '口頭で説明して実演', '30分', '24時間シートを見せながら説明', '10分', '67%削減'],
    ['家族への報告資料', '手書きメモをまとめる', '1時間', '完成版シートを印刷', '10分', '83%削減'],
]

for i, data in enumerate(time_data, 4):
    add_data_row(ws6, i, data)

# 合計時間を計算
total_row = 4 + len(time_data) + 1
ws6.merge_cells(f'A{total_row}:B{total_row}')
ws6[f'A{total_row}'] = '1日あたりの合計削減時間'
ws6[f'A{total_row}'].font = Font(bold=True, size=11)
ws6[f'A{total_row}'].alignment = Alignment(horizontal='right', vertical='center')
ws6[f'A{total_row}'].fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")

ws6[f'C{total_row}'] = '約6時間'
ws6[f'C{total_row}'].font = Font(bold=True, size=11, color="FF0000")
ws6[f'C{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws6[f'C{total_row}'].fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")

ws6[f'D{total_row}'] = '→'
ws6[f'D{total_row}'].font = Font(bold=True, size=14)
ws6[f'D{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws6[f'D{total_row}'].fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")

ws6[f'E{total_row}'] = '約1時間'
ws6[f'E{total_row}'].font = Font(bold=True, size=11, color="01C1AF")
ws6[f'E{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws6[f'E{total_row}'].fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")

ws6[f'F{total_row}'] = '約83%削減'
ws6[f'F{total_row}'].font = Font(bold=True, size=11, color="01C1AF")
ws6[f'F{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws6[f'F{total_row}'].fill = PatternFill(start_color="E0F7F6", end_color="E0F7F6", fill_type="solid")

set_column_widths(ws6, [20, 20, 15, 25, 15, 15])

# ==========================================
# シート7: まとめ
# ==========================================
ws7 = wb.create_sheet("まとめ")
ws7.merge_cells('A1:D1')
ws7['A1'] = '💡 ケア・フィット・サイクルの総合効果まとめ'
ws7['A1'].font = Font(bold=True, size=14, color="01C1AF")
ws7['A1'].alignment = Alignment(horizontal='center', vertical='center')

summary_headers = ['対象者', '主なメリット', '具体的な効果', '実現する価値']
add_header(ws7, 3, summary_headers, header_fill, header_font)

summary_data = [
    ['新人介護士', '覚えることが減る', '24時間シートに全部書いてある\n迷わず実行できる\nベテランのコツが学べる', '安心してケアに集中\n成長スピードが向上'],
    ['ベテラン介護士', '工夫を共有できる', '「コツのフィット」で暗黙知を明文化\n提案が通りやすくなる\n新人育成が楽になる', '経験が活かされる\nやりがいが向上'],
    ['リーダー・ケアマネ', '判断材料が揃う', '記録と気づきが集約\n事務時間が80%削減\nスタッフ育成が進む', '管理業務の効率化\nケアの質が向上'],
    ['夜勤・パート', '情報不足が解消', 'いつでも最新情報を確認\nAI相談窓口でサポート\n申し送りが楽', '孤立感がなくなる\n安心して働ける'],
    ['入居者', '生活リズムが尊重される', '無理な起床時刻がなくなる\n好みが反映される\nどのスタッフでも同じケア', 'QOL（生活の質）向上\n尊厳が守られる'],
    ['施設全体', '属人化が解消', 'ベテランの知恵が共有\n新人育成コストが削減\nケアの質が標準化', '持続可能な運営\n職員の定着率向上'],
]

for i, data in enumerate(summary_data, 4):
    cell = ws7.cell(row=i, column=1, value=data[0])
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    
    for j in range(1, 4):
        cell = ws7.cell(row=i, column=j+1, value=data[j])
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border
        ws7.row_dimensions[i].height = 60

set_column_widths(ws7, [15, 20, 40, 25])

# Excelファイル保存
output_path = '/home/user/webapp/ケアフィットサイクル_課題解決マップ.xlsx'
wb.save(output_path)
print(f'✅ Excelファイルを作成しました: {output_path}')
