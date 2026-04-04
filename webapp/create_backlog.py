#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ケア・フィット・サイクル開発バックログ生成スクリプト
プロダクトバックログ、スプリントバックログ、ユーザーストーリーマップを生成
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def create_backlog_excel():
    wb = Workbook()
    
    # デフォルトシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================
    # 1. プロダクトバックログ（全体要件）
    # ========================================
    ws_product = wb.create_sheet("プロダクトバックログ")
    
    # ヘッダー設定
    headers_product = ["優先度", "Epic", "ユーザーストーリー", "受け入れ条件", "Story Point", "担当Sprint", "ステータス", "備考"]
    ws_product.append(headers_product)
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="01C1AF", end_color="01C1AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws_product[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # プロダクトバックログデータ
    product_backlog = [
        # Epic 1: 基本データ管理
        ["P0", "基本データ管理", "入居者として、基本情報を登録したい", "名前・要介護度・好きなこと・今日のねがいを登録できる", "3", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "基本データ管理", "入居者として、プロフィール写真を表示したい", "丸い画像でハートアイコン付き、温かみのあるUI", "2", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "基本データ管理", "入居者として、14日間の進捗を可視化したい", "Day 1-14のカウント、フェーズ表示（初期/実態ログ/フィット提案/確定）", "2", "Sprint 1", "完了", "MVP実装済み"],
        
        # Epic 2: 24時間シート管理
        ["P0", "24時間シート管理", "介護士として、24時間シートを閲覧したい", "時刻・活動・詳細・ステータスを表示、3カラムレイアウトの左側", "3", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "24時間シート管理", "ケアマネとして、暫定24時間シートを自動生成したい", "アセスメントシートから暫定プランを自動生成（Day 1）", "8", "Sprint 2", "未着手", "AI機能"],
        ["P0", "24時間シート管理", "リーダーとして、24時間シートを編集したい", "時刻・活動・詳細を直接編集、ステータス更新", "5", "Sprint 2", "未着手", "編集機能"],
        ["P1", "24時間シート管理", "介護士として、印刷用フォーマットで出力したい", "家族報告用のPDF/印刷レイアウト", "3", "Sprint 3", "未着手", "帳票機能"],
        
        # Epic 3: ケース記録管理
        ["P0", "ケース記録管理", "介護士として、ケア記録をタイムラインで閲覧したい", "時刻・スタッフ・内容・タグをタイムライン表示", "3", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "ケース記録管理", "介護士として、その場で記録を追加したい", "テキスト入力+送信ボタン、スタッフ情報を自動記録", "2", "Sprint 1", "完了", "MVP実装済み"],
        ["P1", "ケース記録管理", "介護士として、記録を検索・フィルタしたい", "日付・スタッフ・タグで検索・絞り込み", "5", "Sprint 3", "未着手", "検索機能"],
        ["P1", "ケース記録管理", "リーダーとして、記録を編集・削除したい", "誤記入の修正、削除履歴の保存", "3", "Sprint 3", "未着手", "管理機能"],
        
        # Epic 4: デジタル付せん（気づき管理）
        ["P0", "デジタル付せん", "介護士として、気づきを投稿したい", "カテゴリ選択・関連時刻・内容入力、右カラムに目立つボタン", "3", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "デジタル付せん", "介護士として、投稿した気づきを確認したい", "気づきのストックに表示、スタッフ名・日時・内容", "2", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "デジタル付せん", "リーダーとして、付せんを採用/保留したい", "「とりいれる」「まだそのまま」ボタン、24時間シートへ反映", "3", "Sprint 1", "完了", "MVP実装済み"],
        ["P0", "デジタル付せん", "システムとして、時間のフィットを自動検出したい", "起床・就寝等の差が30分以上3日以上続くと提案", "8", "Sprint 2", "未着手", "AI分析"],
        ["P0", "デジタル付せん", "システムとして、好みのフィットを自動検出したい", "記録から感情キーワードを抽出（例：お茶 vs コーヒー）", "8", "Sprint 2", "未着手", "AI分析"],
        ["P0", "デジタル付せん", "システムとして、コツのフィットを自動検出したい", "ベテラン記録から具体的介助手順・声掛けを抽出", "8", "Sprint 2", "未着手", "AI分析"],
        
        # Epic 5: AI分析エンジン
        ["P1", "AI分析エンジン", "システムとして、Gemini APIと連携したい", "Gemini 2.5 Flash/Pro統合、APIキーをCloudflare Secretsで管理", "5", "Sprint 2", "未着手", "基盤構築"],
        ["P1", "AI分析エンジン", "システムとして、計画と実績の乖離を分析したい", "24時間シートとケース記録を照合、乖離パターンを検出", "13", "Sprint 2", "未着手", "コア機能"],
        ["P1", "AI分析エンジン", "システムとして、ベテランのコツを抽出したい", "経験5年以上のスタッフ記録から具体的なノウハウを抽出", "8", "Sprint 2", "未着手", "AI分析"],
        
        # Epic 6: 音声入力
        ["P2", "音声入力", "介護士として、ケア中に音声で記録したい", "Web Speech API統合、音声→テキスト変換", "8", "Sprint 4", "未着手", "拡張機能"],
        ["P2", "音声入力", "夜勤専従として、夜間に音声で記録したい", "薄暗い環境でも使える大きなボタン、音声認識精度の向上", "5", "Sprint 4", "未着手", "拡張機能"],
        
        # Epic 7: AI相談窓口
        ["P2", "AI相談窓口", "新人として、困ったときにAIへ相談したい", "チャット形式、過去記録を読み込んだRAG対応", "13", "Sprint 5", "未着手", "拡張機能"],
        ["P2", "AI相談窓口", "夜勤専従として、夜間の不穏時対応をAIに相談したい", "緊急度判断、過去の対応事例を提示", "8", "Sprint 5", "未着手", "拡張機能"],
        
        # Epic 8: 認証・権限管理
        ["P1", "認証・権限管理", "スタッフとして、ログインしたい", "メールアドレス/パスワード認証、Firebase Authまたは独自実装", "5", "Sprint 3", "未着手", "基盤構築"],
        ["P1", "認証・権限管理", "管理者として、スタッフ権限を管理したい", "新人/ベテラン/リーダー/ケアマネ/看護師の権限設定", "5", "Sprint 3", "未着手", "管理機能"],
        ["P2", "認証・権限管理", "リーダーとして、編集履歴を確認したい", "誰がいつ何を編集したかのログ表示", "3", "Sprint 4", "未着手", "監査機能"],
        
        # Epic 9: 多施設対応
        ["P3", "多施設対応", "事業者として、複数施設を管理したい", "施設ごとのデータ分離、マルチテナント対応", "13", "Sprint 6", "未着手", "将来対応"],
        ["P3", "多施設対応", "事業者として、施設間でベストプラクティスを共有したい", "匿名化された成功事例の共有機能", "8", "Sprint 6", "未着手", "将来対応"],
        
        # Epic 10: パフォーマンス・UX
        ["P1", "パフォーマンス", "ユーザーとして、快適に操作したい", "ページ読み込み3秒以内、API応答500ms以内", "5", "Sprint 3", "未着手", "最適化"],
        ["P2", "パフォーマンス", "ユーザーとして、オフラインでも閲覧したい", "Service Worker、オフライン時のデータキャッシュ", "8", "Sprint 5", "未着手", "拡張機能"],
        ["P1", "UX", "ユーザーとして、スマホで快適に操作したい", "レスポンシブデザイン、タッチ操作最適化", "5", "Sprint 1", "完了", "MVP実装済み"],
    ]
    
    for row_data in product_backlog:
        ws_product.append(row_data)
    
    # 列幅調整
    ws_product.column_dimensions['A'].width = 8
    ws_product.column_dimensions['B'].width = 18
    ws_product.column_dimensions['C'].width = 45
    ws_product.column_dimensions['D'].width = 50
    ws_product.column_dimensions['E'].width = 10
    ws_product.column_dimensions['F'].width = 12
    ws_product.column_dimensions['G'].width = 12
    ws_product.column_dimensions['H'].width = 20
    
    # セルのスタイル設定（優先度に応じた色分け）
    for row_idx, row in enumerate(ws_product.iter_rows(min_row=2), start=2):
        priority = row[0].value
        if priority == "P0":
            fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")  # 赤系
        elif priority == "P1":
            fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # オレンジ系
        elif priority == "P2":
            fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")  # 黄色系
        else:
            fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")  # グレー系
        
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========================================
    # 2. スプリントバックログ（Sprint 1-6）
    # ========================================
    
    sprint_data = {
        "Sprint 1（MVP）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S1-001", "入居者情報登録", "D1データベース設計（residents テーブル）", "Backend", "4", "完了", ""],
            ["S1-002", "入居者情報登録", "入居者CRUD API実装", "Backend", "6", "完了", ""],
            ["S1-003", "入居者情報登録", "プロフィール表示UI実装", "Frontend", "8", "完了", ""],
            ["S1-004", "入居者情報登録", "プロフィール画像表示機能", "Frontend", "4", "完了", ""],
            ["S1-005", "24時間シート閲覧", "D1データベース設計（care_plans テーブル）", "Backend", "4", "完了", ""],
            ["S1-006", "24時間シート閲覧", "24時間シート取得API実装", "Backend", "4", "完了", ""],
            ["S1-007", "24時間シート閲覧", "24時間シート表示UI実装", "Frontend", "8", "完了", ""],
            ["S1-008", "ケース記録管理", "D1データベース設計（case_records テーブル）", "Backend", "4", "完了", ""],
            ["S1-009", "ケース記録管理", "ケース記録CRUD API実装", "Backend", "6", "完了", ""],
            ["S1-010", "ケース記録管理", "ケース記録タイムライン表示UI", "Frontend", "8", "完了", ""],
            ["S1-011", "ケース記録管理", "記録追加フォーム実装", "Frontend", "4", "完了", ""],
            ["S1-012", "デジタル付せん", "D1データベース設計（sticky_notes テーブル）", "Backend", "4", "完了", ""],
            ["S1-013", "デジタル付せん", "付せんCRUD API実装", "Backend", "6", "完了", ""],
            ["S1-014", "デジタル付せん", "気づき投稿UI実装", "Frontend", "8", "完了", ""],
            ["S1-015", "デジタル付せん", "付せん表示・操作UI実装", "Frontend", "8", "完了", ""],
            ["S1-016", "進捗トラッキング", "Day 1-14カウント表示", "Frontend", "4", "完了", ""],
            ["S1-017", "レスポンシブデザイン", "3カラムレイアウト実装", "Frontend", "6", "完了", ""],
            ["S1-018", "レスポンシブデザイン", "スマホ対応（タブレット・モバイル）", "Frontend", "6", "完了", ""],
            ["S1-019", "デプロイ", "Cloudflare Pages設定", "DevOps", "2", "完了", ""],
            ["S1-020", "デプロイ", "D1データベース本番環境構築", "DevOps", "2", "完了", ""],
        ],
        
        "Sprint 2（AI基盤）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S2-001", "AI基盤構築", "Gemini API統合設計", "Backend", "8", "未着手", "API選定・認証設計"],
            ["S2-002", "AI基盤構築", "Cloudflare Secrets設定", "DevOps", "2", "未着手", "API Key管理"],
            ["S2-003", "AI基盤構築", "Gemini API呼び出しラッパー実装", "Backend", "8", "未着手", ""],
            ["S2-004", "時間のフィット", "計画と実績の時刻差分析ロジック", "Backend", "12", "未着手", "30分以上×3日以上"],
            ["S2-005", "時間のフィット", "時間フィット提案生成API", "Backend", "8", "未着手", ""],
            ["S2-006", "時間のフィット", "時間フィット付せん表示UI", "Frontend", "6", "未着手", ""],
            ["S2-007", "好みのフィット", "感情キーワード抽出ロジック", "Backend", "12", "未着手", "NLP処理"],
            ["S2-008", "好みのフィット", "好みフィット提案生成API", "Backend", "8", "未着手", ""],
            ["S2-009", "好みのフィット", "好みフィット付せん表示UI", "Frontend", "6", "未着手", ""],
            ["S2-010", "コツのフィット", "ベテラン記録抽出ロジック", "Backend", "12", "未着手", "経験5年以上"],
            ["S2-011", "コツのフィット", "コツフィット提案生成API", "Backend", "8", "未着手", ""],
            ["S2-012", "コツのフィット", "コツフィット付せん表示UI", "Frontend", "6", "未着手", ""],
            ["S2-013", "暫定24時間シート自動生成", "アセスメントシート入力フォーム", "Frontend", "8", "未着手", ""],
            ["S2-014", "暫定24時間シート自動生成", "AI自動生成ロジック実装", "Backend", "16", "未着手", "プロンプト設計"],
            ["S2-015", "24時間シート編集", "編集モードUI実装", "Frontend", "8", "未着手", ""],
            ["S2-016", "24時間シート編集", "更新API実装", "Backend", "4", "未着手", ""],
            ["S2-017", "テスト", "AI提案精度検証", "QA", "8", "未着手", ""],
        ],
        
        "Sprint 3（認証・管理）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S3-001", "スタッフ認証", "認証方式選定（Firebase Auth or 独自）", "Backend", "4", "未着手", ""],
            ["S3-002", "スタッフ認証", "ログインAPI実装", "Backend", "8", "未着手", ""],
            ["S3-003", "スタッフ認証", "ログインUI実装", "Frontend", "6", "未着手", ""],
            ["S3-004", "スタッフ認証", "セッション管理実装", "Backend", "6", "未着手", "JWT or Cookie"],
            ["S3-005", "権限管理", "ロール定義（新人/ベテラン/リーダー等）", "Backend", "4", "未着手", ""],
            ["S3-006", "権限管理", "権限チェックミドルウェア実装", "Backend", "6", "未着手", ""],
            ["S3-007", "権限管理", "権限管理UI実装", "Frontend", "8", "未着手", "管理者画面"],
            ["S3-008", "記録検索", "検索API実装（日付・スタッフ・タグ）", "Backend", "8", "未着手", ""],
            ["S3-009", "記録検索", "検索UI実装", "Frontend", "6", "未着手", ""],
            ["S3-010", "記録編集・削除", "編集・削除API実装", "Backend", "6", "未着手", ""],
            ["S3-011", "記録編集・削除", "編集・削除UI実装", "Frontend", "6", "未着手", ""],
            ["S3-012", "印刷フォーマット", "PDF生成ライブラリ選定", "Backend", "2", "未着手", ""],
            ["S3-013", "印刷フォーマット", "PDF生成API実装", "Backend", "8", "未着手", ""],
            ["S3-014", "印刷フォーマット", "印刷レイアウトUI実装", "Frontend", "6", "未着手", ""],
            ["S3-015", "パフォーマンス最適化", "API応答速度計測", "DevOps", "4", "未着手", ""],
            ["S3-016", "パフォーマンス最適化", "データベースインデックス最適化", "Backend", "4", "未着手", ""],
            ["S3-017", "パフォーマンス最適化", "フロントエンドバンドルサイズ削減", "Frontend", "4", "未着手", ""],
        ],
        
        "Sprint 4（拡張機能）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S4-001", "音声入力", "Web Speech API調査・検証", "Frontend", "4", "未着手", ""],
            ["S4-002", "音声入力", "音声→テキスト変換実装", "Frontend", "8", "未着手", ""],
            ["S4-003", "音声入力", "音声入力UI実装", "Frontend", "6", "未着手", "大きなボタン"],
            ["S4-004", "音声入力", "音声認識精度向上（ノイズ除去）", "Frontend", "8", "未着手", ""],
            ["S4-005", "編集履歴", "履歴テーブル設計", "Backend", "4", "未着手", ""],
            ["S4-006", "編集履歴", "履歴記録ロジック実装", "Backend", "6", "未着手", ""],
            ["S4-007", "編集履歴", "履歴表示UI実装", "Frontend", "6", "未着手", ""],
            ["S4-008", "通知機能", "通知テーブル設計", "Backend", "4", "未着手", ""],
            ["S4-009", "通知機能", "通知API実装", "Backend", "6", "未着手", ""],
            ["S4-010", "通知機能", "通知表示UI実装", "Frontend", "6", "未着手", "ベル・バッジ"],
        ],
        
        "Sprint 5（AI相談窓口）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S5-001", "AI相談窓口", "RAG（Retrieval-Augmented Generation）設計", "Backend", "8", "未着手", ""],
            ["S5-002", "AI相談窓口", "ベクトルDB選定・構築", "Backend", "8", "未着手", "Cloudflare Vectorize"],
            ["S5-003", "AI相談窓口", "記録のベクトル化実装", "Backend", "8", "未着手", ""],
            ["S5-004", "AI相談窓口", "チャットAPI実装", "Backend", "12", "未着手", ""],
            ["S5-005", "AI相談窓口", "チャットUI実装", "Frontend", "12", "未着手", ""],
            ["S5-006", "AI相談窓口", "緊急度判断ロジック実装", "Backend", "8", "未着手", ""],
            ["S5-007", "AI相談窓口", "過去事例提示機能", "Backend", "8", "未着手", ""],
            ["S5-008", "オフライン対応", "Service Worker実装", "Frontend", "8", "未着手", ""],
            ["S5-009", "オフライン対応", "データキャッシュ戦略設計", "Frontend", "6", "未着手", ""],
            ["S5-010", "オフライン対応", "同期ロジック実装", "Backend", "8", "未着手", ""],
        ],
        
        "Sprint 6（多施設対応）": [
            ["タスクID", "ユーザーストーリー", "タスク", "担当者", "工数(h)", "ステータス", "備考"],
            ["S6-001", "マルチテナント", "テナント分離設計", "Backend", "8", "未着手", ""],
            ["S6-002", "マルチテナント", "施設マスターテーブル設計", "Backend", "4", "未着手", ""],
            ["S6-003", "マルチテナント", "施設切り替えUI実装", "Frontend", "6", "未着手", ""],
            ["S6-004", "ベストプラクティス共有", "匿名化ロジック実装", "Backend", "8", "未着手", ""],
            ["S6-005", "ベストプラクティス共有", "事例共有API実装", "Backend", "8", "未着手", ""],
            ["S6-006", "ベストプラクティス共有", "事例閲覧UI実装", "Frontend", "8", "未着手", ""],
            ["S6-007", "施設管理", "施設CRUD API実装", "Backend", "6", "未着手", ""],
            ["S6-008", "施設管理", "施設管理画面実装", "Frontend", "8", "未着手", ""],
        ],
    }
    
    for sprint_name, tasks in sprint_data.items():
        ws_sprint = wb.create_sheet(sprint_name)
        
        # ヘッダー行
        ws_sprint.append(tasks[0])
        for cell in ws_sprint[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # データ行
        for task_row in tasks[1:]:
            ws_sprint.append(task_row)
        
        # 列幅調整
        ws_sprint.column_dimensions['A'].width = 12
        ws_sprint.column_dimensions['B'].width = 30
        ws_sprint.column_dimensions['C'].width = 40
        ws_sprint.column_dimensions['D'].width = 12
        ws_sprint.column_dimensions['E'].width = 10
        ws_sprint.column_dimensions['F'].width = 12
        ws_sprint.column_dimensions['G'].width = 25
        
        # セルのスタイル設定
        for row in ws_sprint.iter_rows(min_row=2):
            status = row[5].value
            if status == "完了":
                fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")  # 緑系
            elif status == "進行中":
                fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # オレンジ系
            else:
                fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # グレー系
            
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    # ========================================
    # 3. ユーザーストーリーマップ
    # ========================================
    ws_user_story = wb.create_sheet("ユーザーストーリーマップ")
    
    headers_user_story = ["職種", "フェーズ", "活動", "ユーザーストーリー", "画面/機能", "優先度"]
    ws_user_story.append(headers_user_story)
    
    for cell in ws_user_story[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    user_story_map = [
        # 新人介護士
        ["新人介護士", "Day 1", "初期確認", "暫定24時間シートを確認したい", "24時間シート閲覧", "P0"],
        ["新人介護士", "Day 2-7", "ケア実施", "朝の起床介助を24時間シートを見ながら実施したい", "24時間シート閲覧", "P0"],
        ["新人介護士", "Day 2-7", "記録", "ケア後すぐに記録を追加したい", "ケース記録追加", "P0"],
        ["新人介護士", "Day 2-7", "記録", "音声で記録を追加したい（手が離せない時）", "音声入力", "P2"],
        ["新人介護士", "Day 8-13", "学習", "ベテランのコツを学びたい", "コツのフィット閲覧", "P1"],
        ["新人介護士", "Day 8-13", "相談", "困ったときにAIに相談したい", "AI相談窓口", "P2"],
        
        # ベテラン介護士
        ["ベテラン介護士", "Day 2-7", "気づき", "岡田様の好みの変化に気づいたので記録したい", "気づき投稿", "P0"],
        ["ベテラン介護士", "Day 2-7", "実験", "ティータイムの時間を変えて反応を見たい", "ケース記録追加", "P0"],
        ["ベテラン介護士", "Day 8-13", "共有", "工夫した介助方法をチームに共有したい", "気づき投稿（コツのフィット）", "P0"],
        ["ベテラン介護士", "Day 8-13", "提案", "時間のズレを調整する提案をしたい", "気づき投稿（時間のフィット）", "P0"],
        
        # ケアマネージャー/リーダー
        ["ケアマネ", "Day 1", "計画作成", "アセスメントシートから暫定24時間シートを自動生成したい", "暫定24時間シート自動生成", "P0"],
        ["ケアマネ", "Day 2-7", "確認", "朝礼で全スタッフの気づきを確認したい", "気づきのストック閲覧", "P0"],
        ["ケアマネ", "Day 8-13", "判断", "付せんを採用/保留したい", "付せん操作（とりいれる/まだそのまま）", "P0"],
        ["ケアマネ", "Day 14", "完成", "完成版24時間シートを確認・承認したい", "24時間シート閲覧・編集", "P0"],
        ["ケアマネ", "Day 14", "報告", "家族報告用資料を印刷したい", "印刷フォーマット出力", "P1"],
        
        # 看護師
        ["看護師", "Day 2-7", "確認", "朝の健康チェック前に前夜の記録を確認したい", "ケース記録閲覧（検索・フィルタ）", "P1"],
        ["看護師", "Day 2-7", "記録", "服薬介助時の異変を記録したい", "ケース記録追加", "P0"],
        ["看護師", "Day 8-13", "判断", "体調変化の付せんを医療的に判断したい", "付せん操作", "P0"],
        
        # 夜勤・夜勤専従
        ["夜勤専従", "Day 2-7", "引き継ぎ", "夜勤前に日中の記録を確認したい", "ケース記録閲覧", "P0"],
        ["夜勤専従", "Day 2-7", "記録", "夜間の見回り時にスマホから記録したい", "ケース記録追加（スマホ対応）", "P0"],
        ["夜勤専従", "Day 2-7", "相談", "不穏時対応をAIに相談したい", "AI相談窓口", "P2"],
        
        # パート/非常勤
        ["パート", "Day 2-7", "確認", "出勤時に過去7日分の記録を一括確認したい", "ケース記録閲覧（日付フィルタ）", "P1"],
        ["パート", "Day 8-13", "提案", "気づいたコツを投稿したい", "気づき投稿", "P0"],
    ]
    
    for row_data in user_story_map:
        ws_user_story.append(row_data)
    
    # 列幅調整
    ws_user_story.column_dimensions['A'].width = 15
    ws_user_story.column_dimensions['B'].width = 12
    ws_user_story.column_dimensions['C'].width = 15
    ws_user_story.column_dimensions['D'].width = 50
    ws_user_story.column_dimensions['E'].width = 30
    ws_user_story.column_dimensions['F'].width = 8
    
    # セルのスタイル設定
    for row in ws_user_story.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========================================
    # 4. 技術仕様書
    # ========================================
    ws_tech = wb.create_sheet("技術仕様書")
    
    headers_tech = ["カテゴリ", "項目", "詳細", "備考"]
    ws_tech.append(headers_tech)
    
    for cell in ws_tech[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    tech_spec = [
        ["アーキテクチャ", "全体構成", "Cloudflare Pages + Hono + D1 + Gemini API", "エッジコンピューティング"],
        ["アーキテクチャ", "フロントエンド", "React (CDN) + TailwindCSS (CDN)", "軽量・高速"],
        ["アーキテクチャ", "バックエンド", "Hono (TypeScript)", "Cloudflare Workers"],
        ["アーキテクチャ", "データベース", "Cloudflare D1 (SQLite)", "グローバル分散"],
        ["アーキテクチャ", "AI", "Gemini 2.5 Flash / Pro", "Google Generative AI"],
        ["アーキテクチャ", "認証", "Firebase Auth または 独自実装", "Sprint 3で選定"],
        ["", "", "", ""],
        
        ["データベース設計", "residents", "入居者マスター（id, name, care_level, favorite_things, today_wish, maturation_day, phase）", ""],
        ["データベース設計", "care_plans", "24時間シート（id, resident_id, time, activity, details, status, display_order）", ""],
        ["データベース設計", "case_records", "ケース記録（id, resident_id, staff_id, record_time, content, tag, record_type）", ""],
        ["データベース設計", "sticky_notes", "付せん（id, resident_id, staff_id, note_type, related_time, title, content, source, status）", ""],
        ["データベース設計", "staff", "スタッフマスター（id, name, experience_years, role）", ""],
        ["", "", "", ""],
        
        ["API設計", "GET /api/residents/:id", "入居者情報取得", ""],
        ["API設計", "PUT /api/residents/:id", "入居者情報更新", ""],
        ["API設計", "GET /api/residents/:id/care-plans", "24時間シート取得", ""],
        ["API設計", "PUT /api/care-plans/:id", "24時間シート更新", ""],
        ["API設計", "GET /api/residents/:id/case-records", "ケース記録取得", ""],
        ["API設計", "POST /api/case-records", "ケース記録追加", ""],
        ["API設計", "GET /api/residents/:id/sticky-notes", "付せん取得", ""],
        ["API設計", "POST /api/sticky-notes", "付せん追加", ""],
        ["API設計", "PUT /api/sticky-notes/:id", "付せん更新", ""],
        ["API設計", "GET /api/staff", "スタッフ一覧取得", ""],
        ["", "", "", ""],
        
        ["セキュリティ", "API認証", "JWT または Cookie-based Session", "Sprint 3で実装"],
        ["セキュリティ", "APIキー管理", "Cloudflare Secrets", "Gemini API Key"],
        ["セキュリティ", "CORS", "API エンドポイントで有効化", ""],
        ["セキュリティ", "データ暗号化", "HTTPS強制", ""],
        ["", "", "", ""],
        
        ["パフォーマンス", "ページ読み込み", "3秒以内（目標）", ""],
        ["パフォーマンス", "API応答", "500ms以内（目標）", ""],
        ["パフォーマンス", "データベース", "インデックス最適化", "Sprint 3"],
        ["パフォーマンス", "バンドルサイズ", "Worker: 30KB以内", "現在: 28KB"],
        ["", "", "", ""],
        
        ["デプロイ", "ローカル開発", "npm run dev（PM2 + wrangler pages dev）", ""],
        ["デプロイ", "本番デプロイ", "npm run deploy:prod", ""],
        ["デプロイ", "環境変数", ".dev.vars（ローカル）、Cloudflare Secrets（本番）", ""],
        ["", "", "", ""],
        
        ["テスト", "単体テスト", "Vitest", "Sprint 2以降"],
        ["テスト", "E2Eテスト", "Playwright", "Sprint 3以降"],
        ["テスト", "AI精度検証", "手動テスト + メトリクス測定", "Sprint 2"],
    ]
    
    for row_data in tech_spec:
        ws_tech.append(row_data)
    
    # 列幅調整
    ws_tech.column_dimensions['A'].width = 20
    ws_tech.column_dimensions['B'].width = 30
    ws_tech.column_dimensions['C'].width = 60
    ws_tech.column_dimensions['D'].width = 20
    
    # セルのスタイル設定
    for row in ws_tech.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========================================
    # 5. 開発スケジュール
    # ========================================
    ws_schedule = wb.create_sheet("開発スケジュール")
    
    headers_schedule = ["Sprint", "期間", "目標", "完了条件", "リスク", "工数合計"]
    ws_schedule.append(headers_schedule)
    
    for cell in ws_schedule[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    schedule_data = [
        ["Sprint 1（MVP）", "Week 1-2", "基本機能の実装とデプロイ", "入居者管理・24時間シート・ケース記録・付せんが動作する", "なし（実装済み）", "100h"],
        ["Sprint 2（AI基盤）", "Week 3-4", "AI分析エンジンの構築", "時間・好み・コツのフィットが自動提案される", "Gemini API統合の複雑さ", "120h"],
        ["Sprint 3（認証・管理）", "Week 5-6", "認証機能と管理機能の追加", "スタッフログイン・権限管理・検索・印刷が動作する", "認証方式の選定に時間がかかる可能性", "90h"],
        ["Sprint 4（拡張機能）", "Week 7-8", "音声入力と履歴管理", "音声で記録追加・編集履歴が確認できる", "Web Speech APIの精度", "60h"],
        ["Sprint 5（AI相談窓口）", "Week 9-10", "AI相談窓口とオフライン対応", "AIチャットが動作・オフラインでも閲覧可能", "RAG実装の複雑さ", "80h"],
        ["Sprint 6（多施設対応）", "Week 11-12", "マルチテナント対応", "複数施設を管理・ベストプラクティス共有", "データ分離の設計", "60h"],
    ]
    
    for row_data in schedule_data:
        ws_schedule.append(row_data)
    
    # 列幅調整
    ws_schedule.column_dimensions['A'].width = 18
    ws_schedule.column_dimensions['B'].width = 15
    ws_schedule.column_dimensions['C'].width = 30
    ws_schedule.column_dimensions['D'].width = 50
    ws_schedule.column_dimensions['E'].width = 40
    ws_schedule.column_dimensions['F'].width = 12
    
    # セルのスタイル設定
    for row in ws_schedule.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # ========================================
    # ファイル保存
    # ========================================
    output_path = "/home/user/webapp/ケアフィットサイクル_開発バックログ.xlsx"
    wb.save(output_path)
    print(f"✅ バックログファイルを作成しました: {output_path}")

if __name__ == "__main__":
    create_backlog_excel()
